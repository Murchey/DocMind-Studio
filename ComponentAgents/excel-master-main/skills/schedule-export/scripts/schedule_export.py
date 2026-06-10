"""
Schedule Export - 将排课结果导出为 Excel 文件
支持按班级、教师、教室等维度生成课表
"""

import json
import os
from pathlib import Path
from typing import Dict, List, Any
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ScheduleExporter:
    """排课结果导出器"""
    
    # 时间槽定义
    TIME_SLOTS = [
        {"period": 1, "time": "08:00-08:45", "label": "第1节"},
        {"period": 2, "time": "08:55-09:40", "label": "第2节"},
        {"period": 3, "time": "10:00-10:45", "label": "第3节"},
        {"period": 4, "time": "10:55-11:40", "label": "第4节"},
        {"period": 5, "time": "14:00-14:45", "label": "第5节"},
        {"period": 6, "time": "14:55-15:40", "label": "第6节"},
        {"period": 7, "time": "16:00-16:45", "label": "第7节"},
        {"period": 8, "time": "16:55-17:40", "label": "第8节"},
    ]
    
    DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    DAY_LABELS = {"Monday": "周一", "Tuesday": "周二", "Wednesday": "周三", "Thursday": "周四", "Friday": "周五"}
    
    def __init__(self):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required. Install it with: pip install openpyxl")
    
    def export_all(self, schedule_path: str, output_dir: str):
        """
        导出所有视图的课表
        
        Args:
            schedule_path: 排课结果 JSON 文件路径
            output_dir: 输出目录
        """
        # 创建输出目录
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        
        # 导出各种视图
        self.export_by_class(schedule_path, os.path.join(output_dir, "schedule_by_class.xlsx"))
        self.export_by_teacher(schedule_path, os.path.join(output_dir, "schedule_by_teacher.xlsx"))
        self.export_by_room(schedule_path, os.path.join(output_dir, "schedule_by_room.xlsx"))
        
        print(f"Exported schedules to {output_dir}")
    
    def export_by_class(self, schedule_path: str, output_path: str):
        """按班级导出课表"""
        with open(schedule_path, 'r', encoding='utf-8') as f:
            schedule = json.load(f)
        
        assignments = schedule.get("assignments", [])
        
        # 按班级分组
        class_assignments = defaultdict(list)
        for a in assignments:
            class_assignments[a["class_name"]].append(a)
        
        # 创建工作簿
        wb = Workbook()
        
        # 为每个班级创建一个工作表
        for class_name, class_assigns in class_assignments.items():
            ws = wb.create_sheet(title=class_name[:31])  # Excel 限制 31 字符
            self._fill_class_sheet(ws, class_name, class_assigns)
        
        # 删除默认工作表
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # 保存
        wb.save(output_path)
        print(f"Exported class schedule to {output_path}")
    
    def export_by_teacher(self, schedule_path: str, output_path: str):
        """按教师导出课表"""
        with open(schedule_path, 'r', encoding='utf-8') as f:
            schedule = json.load(f)
        
        assignments = schedule.get("assignments", [])
        
        # 按教师分组
        teacher_assignments = defaultdict(list)
        for a in assignments:
            teacher_assignments[a["teacher_name"]].append(a)
        
        # 创建工作簿
        wb = Workbook()
        
        # 为每个教师创建一个工作表
        for teacher_name, teacher_assigns in teacher_assignments.items():
            ws = wb.create_sheet(title=teacher_name[:31])
            self._fill_teacher_sheet(ws, teacher_name, teacher_assigns)
        
        # 删除默认工作表
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # 保存
        wb.save(output_path)
        print(f"Exported teacher schedule to {output_path}")
    
    def export_by_room(self, schedule_path: str, output_path: str):
        """按教室导出课表"""
        with open(schedule_path, 'r', encoding='utf-8') as f:
            schedule = json.load(f)
        
        assignments = schedule.get("assignments", [])
        
        # 按教室分组
        room_assignments = defaultdict(list)
        for a in assignments:
            room_assignments[a["room_name"]].append(a)
        
        # 创建工作簿
        wb = Workbook()
        
        # 为每个教室创建一个工作表
        for room_name, room_assigns in room_assignments.items():
            ws = wb.create_sheet(title=room_name[:31])
            self._fill_room_sheet(ws, room_name, room_assigns)
        
        # 删除默认工作表
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # 保存
        wb.save(output_path)
        print(f"Exported room schedule to {output_path}")
    
    def _fill_class_sheet(self, ws, class_name: str, assignments: List[Dict]):
        """填充班级课表"""
        # 设置标题
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"{class_name} 课程表"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # 设置表头
        headers = ["时间"] + [self.DAY_LABELS[d] for d in self.DAYS]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 填充时间表
        for row_idx, slot in enumerate(self.TIME_SLOTS, 4):
            # 时间列
            cell = ws.cell(row=row_idx, column=1)
            cell.value = f"{slot['label']}\n{slot['time']}"
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            
            # 星期列
            for col_idx, day in enumerate(self.DAYS, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # 查找对应的课程
                for a in assignments:
                    if a["day"] == day and a["period"] == slot["period"]:
                        cell.value = f"{a['course_name']}\n{a['teacher_name']}\n{a['room_name']}"
                        break
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        for col in range(2, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # 设置行高
        for row in range(3, 12):
            ws.row_dimensions[row].height = 40
    
    def _fill_teacher_sheet(self, ws, teacher_name: str, assignments: List[Dict]):
        """填充教师课表"""
        # 设置标题
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"{teacher_name} 课程表"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # 设置表头
        headers = ["时间"] + [self.DAY_LABELS[d] for d in self.DAYS]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 填充时间表
        for row_idx, slot in enumerate(self.TIME_SLOTS, 4):
            # 时间列
            cell = ws.cell(row=row_idx, column=1)
            cell.value = f"{slot['label']}\n{slot['time']}"
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            
            # 星期列
            for col_idx, day in enumerate(self.DAYS, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # 查找对应的课程
                for a in assignments:
                    if a["day"] == day and a["period"] == slot["period"]:
                        cell.value = f"{a['course_name']}\n{a['class_name']}\n{a['room_name']}"
                        break
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        for col in range(2, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # 设置行高
        for row in range(3, 12):
            ws.row_dimensions[row].height = 40
    
    def _fill_room_sheet(self, ws, room_name: str, assignments: List[Dict]):
        """填充教室课表"""
        # 设置标题
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"{room_name} 使用表"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # 设置表头
        headers = ["时间"] + [self.DAY_LABELS[d] for d in self.DAYS]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 填充时间表
        for row_idx, slot in enumerate(self.TIME_SLOTS, 4):
            # 时间列
            cell = ws.cell(row=row_idx, column=1)
            cell.value = f"{slot['label']}\n{slot['time']}"
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            
            # 星期列
            for col_idx, day in enumerate(self.DAYS, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # 查找对应的课程
                for a in assignments:
                    if a["day"] == day and a["period"] == slot["period"]:
                        cell.value = f"{a['course_name']}\n{a['teacher_name']}\n{a['class_name']}"
                        break
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        for col in range(2, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # 设置行高
        for row in range(3, 12):
            ws.row_dimensions[row].height = 40


def main():
    """命令行入口"""
    import argparse
    
    parser = argparse.ArgumentParser(description='导出排课结果为 Excel')
    parser.add_argument('input', help='排课结果 JSON 文件路径')
    parser.add_argument('output', help='输出目录路径')
    parser.add_argument('--view', choices=['class', 'teacher', 'room', 'all'], default='all', help='导出视图类型')
    
    args = parser.parse_args()
    
    exporter = ScheduleExporter()
    
    if args.view == 'all':
        exporter.export_all(args.input, args.output)
    elif args.view == 'class':
        Path(args.output).mkdir(parents=True, exist_ok=True)
        exporter.export_by_class(args.input, os.path.join(args.output, "schedule_by_class.xlsx"))
    elif args.view == 'teacher':
        Path(args.output).mkdir(parents=True, exist_ok=True)
        exporter.export_by_teacher(args.input, os.path.join(args.output, "schedule_by_teacher.xlsx"))
    elif args.view == 'room':
        Path(args.output).mkdir(parents=True, exist_ok=True)
        exporter.export_by_room(args.input, os.path.join(args.output, "schedule_by_room.xlsx"))


if __name__ == '__main__':
    main()
